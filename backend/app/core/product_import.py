"""Bulk product import — parsing, column mapping, validation and category
matching. Kept separate from the route so it is unit-testable and never touches
the working single-product create/update endpoints.

Design:
- The canonical columns mirror the Product model / ProductCreate schema; the
  importer never invents a new data shape.
- Rows are validated (types + required) and matched to the existing category
  tree (by slug, name EN/BN, or a "A > B" path) — no new taxonomy is created.
- slug is the primary key: an existing slug updates, a new slug creates. SKU
  duplicates (in-file or against a different product) are surfaced as warnings.
- Validation is a pure dry-run (no DB writes); the caller commits separately.
"""
from __future__ import annotations

import csv
import io
import re
from dataclasses import dataclass, field
from typing import Any

from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.models.models import Category

# ── Column spec ──────────────────────────────────────────────────────────────
# type: str | int | float | bool | list ; required only for a *new* product.
FIELD_SPEC: dict[str, dict] = {
    "slug": {"type": "str", "required": True, "aliases": ["url", "handle"]},
    "name_en": {"type": "str", "required": True, "aliases": ["name", "name(english)", "nameen", "title", "title_en"]},
    "name_bn": {"type": "str", "required": True, "aliases": ["name(bangla)", "namebn", "title_bn", "bangla name"]},
    "category": {"type": "str", "required": True, "aliases": ["category_slug", "categoryname", "category name", "category_path", "categorypath"]},
    "price": {"type": "float", "required": True, "aliases": ["mrp", "sell price", "sellingprice"]},
    "original_price": {"type": "float", "aliases": ["compare price", "compareatprice", "old price", "regular price"]},
    "description_en": {"type": "str", "aliases": ["description", "desc_en", "details"]},
    "description_bn": {"type": "str", "aliases": ["desc_bn", "bangla description"]},
    "sku": {"type": "str", "aliases": ["product code", "code", "item code"]},
    "barcode": {"type": "str", "aliases": ["ean", "upc"]},
    "brand": {"type": "str", "aliases": ["manufacturer"]},
    "stock_quantity": {"type": "int", "aliases": ["stock", "qty", "quantity", "inventory"]},
    "low_stock_threshold": {"type": "int", "aliases": ["low stock", "reorder level"]},
    "is_active": {"type": "bool", "aliases": ["active", "published", "status"]},
    "is_featured": {"type": "bool", "aliases": ["featured"]},
    "is_best_seller": {"type": "bool", "aliases": ["best seller", "bestseller"]},
    "badge": {"type": "str", "aliases": ["tag label", "ribbon"]},
    "image_url": {"type": "str", "aliases": ["image", "main image", "thumbnail", "photo"]},
    "images": {"type": "list", "aliases": ["gallery", "additional images", "more images"]},
    "tags": {"type": "list", "aliases": ["keywords", "labels"]},
    "sort_order": {"type": "int", "aliases": ["order", "position"]},
    "weight": {"type": "float", "aliases": ["weight(kg)", "kg"]},
    "delivery_charge": {"type": "float", "aliases": ["shipping", "delivery"]},
    "warranty_info": {"type": "str", "aliases": ["warranty"]},
    "seo_title": {"type": "str", "aliases": []},
    "seo_description": {"type": "str", "aliases": []},
    "seo_keywords": {"type": "str", "aliases": []},
}

# Header (normalised) -> canonical field, built from names + aliases.
def _norm(h: str) -> str:
    return re.sub(r"[\s_\-()]+", "", (h or "").strip().lower())


_HEADER_LOOKUP: dict[str, str] = {}
for _fname, _spec in FIELD_SPEC.items():
    _HEADER_LOOKUP[_norm(_fname)] = _fname
    for _al in _spec.get("aliases", []):
        _HEADER_LOOKUP.setdefault(_norm(_al), _fname)


def auto_map_headers(headers: list[str]) -> dict[str, str | None]:
    """Return {original_header: canonical_field_or_None} by name/alias match."""
    return {h: _HEADER_LOOKUP.get(_norm(h)) for h in headers}


def slugify(text: str) -> str:
    s = re.sub(r"[^a-z0-9]+", "-", (text or "").lower()).strip("-")
    return s or ""


# ── Value coercion ───────────────────────────────────────────────────────────
_TRUE = {"true", "1", "yes", "y", "on", "active", "published", "হ্যাঁ"}
_FALSE = {"false", "0", "no", "n", "off", "inactive", "draft", "না", ""}


def _coerce(field_name: str, raw: Any) -> tuple[Any, str | None]:
    """Coerce a raw cell to the field's type. Returns (value, error)."""
    spec = FIELD_SPEC[field_name]
    t = spec["type"]
    s = "" if raw is None else str(raw).strip()
    if s == "":
        return (None, None)
    try:
        if t == "str":
            return (s, None)
        if t == "int":
            return (int(float(s)), None)  # tolerate "5.0"
        if t == "float":
            return (float(s.replace(",", "")), None)
        if t == "bool":
            low = s.lower()
            if low in _TRUE:
                return (True, None)
            if low in _FALSE:
                return (False, None)
            return (None, f"'{s}' is not a yes/no value")
        if t == "list":
            parts = [p.strip() for p in re.split(r"[,\n|]+", s) if p.strip()]
            return (parts, None)
    except (ValueError, TypeError):
        return (None, f"'{s}' is not a valid {t}")
    return (s, None)


# ── Category matching ────────────────────────────────────────────────────────
@dataclass
class CategoryIndex:
    by_slug: dict[str, Category]
    by_name: dict[str, Category]
    by_id: dict[Any, Category]

    def resolve(self, value: str) -> Category | None:
        """Match a category cell to a node: slug, name (EN/BN), or 'A > B' path
        (the last path segment wins, matched by name or slug)."""
        v = (value or "").strip()
        if not v:
            return None
        if ">" in v or "/" in v or "»" in v:
            v = re.split(r"[>/»]", v)[-1].strip()
        key = v.lower()
        return self.by_slug.get(key) or self.by_name.get(key) or self.by_slug.get(slugify(v))


async def build_category_index(db: AsyncSession) -> CategoryIndex:
    cats = list((await db.execute(
        select(Category).where(Category.is_deleted == False)  # noqa: E712
    )).scalars().all())
    by_slug: dict[str, Category] = {}
    by_name: dict[str, Category] = {}
    by_id: dict[Any, Category] = {}
    for c in cats:
        by_id[c.id] = c
        if c.slug:
            by_slug[c.slug.lower()] = c
        if c.name_en:
            by_name.setdefault(c.name_en.strip().lower(), c)
        if c.name_bn:
            by_name.setdefault(c.name_bn.strip().lower(), c)
    return CategoryIndex(by_slug=by_slug, by_name=by_name, by_id=by_id)


def category_path(cat: Category, index: CategoryIndex) -> str:
    """Root-first 'A > B > C' label for a node."""
    chain: list[str] = []
    seen: set[Any] = set()
    cur: Category | None = cat
    while cur is not None and cur.id not in seen:
        seen.add(cur.id)
        chain.append(cur.name_en or cur.slug)
        cur = index.by_id.get(cur.parent_id) if cur.parent_id else None
    return " > ".join(reversed(chain))


# ── Parsing ──────────────────────────────────────────────────────────────────
def parse_file(filename: str, content: bytes) -> tuple[list[str], list[dict[str, str]]]:
    """Return (headers, rows) from a CSV or XLSX upload. Raises ValueError on a
    format we cannot read."""
    name = (filename or "").lower()
    if name.endswith(".xlsx") or name.endswith(".xlsm"):
        return _parse_xlsx(content)
    # default: CSV (also handles .txt / unknown as CSV)
    text = content.decode("utf-8-sig", errors="replace")
    reader = csv.reader(io.StringIO(text))
    all_rows = [r for r in reader]
    if not all_rows:
        return ([], [])
    headers = [h.strip() for h in all_rows[0]]
    rows: list[dict[str, str]] = []
    for r in all_rows[1:]:
        if not any((c or "").strip() for c in r):
            continue  # skip fully-blank lines
        rows.append({headers[i]: (r[i] if i < len(r) else "") for i in range(len(headers))})
    return (headers, rows)


def _parse_xlsx(content: bytes) -> tuple[list[str], list[dict[str, str]]]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:  # pragma: no cover
        raise ValueError("Excel support is unavailable on the server (openpyxl not installed).") from exc
    wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    ws = wb.active
    rows_iter = ws.iter_rows(values_only=True)
    try:
        header_row = next(rows_iter)
    except StopIteration:
        return ([], [])
    headers = [str(h).strip() if h is not None else "" for h in header_row]
    rows: list[dict[str, str]] = []
    for r in rows_iter:
        if r is None or not any(c is not None and str(c).strip() for c in r):
            continue
        row: dict[str, str] = {}
        for i, h in enumerate(headers):
            if not h:
                continue
            val = r[i] if i < len(r) else None
            row[h] = "" if val is None else str(val).strip()
        rows.append(row)
    wb.close()
    return (headers, rows)


# ── Row validation ───────────────────────────────────────────────────────────
@dataclass
class RowResult:
    row_num: int
    data: dict[str, Any] = field(default_factory=dict)  # coerced canonical fields
    category_id: Any = None
    category_label: str = ""
    action: str = "skip"  # create | update | skip
    errors: list[str] = field(default_factory=list)
    warnings: list[str] = field(default_factory=list)

    def to_preview(self) -> dict:
        return {
            "row": self.row_num,
            "slug": self.data.get("slug", ""),
            "name": self.data.get("name_en", ""),
            "sku": self.data.get("sku", ""),
            "category": self.category_label,
            "action": "skip" if self.errors else self.action,
            "errors": self.errors,
            "warnings": self.warnings,
        }


def validate_rows(
    rows: list[dict[str, str]],
    mapping: dict[str, str | None],
    cat_index: CategoryIndex,
    existing_slugs: set[str],
    existing_sku_to_slug: dict[str, str],
    on_existing: str = "update",
    on_new: str = "create",
) -> list[RowResult]:
    """Pure validation. `mapping` is {original_header: canonical_field|None}.
    `existing_slugs`/`existing_sku_to_slug` come from the DB (case as stored)."""
    # Invert mapping: canonical_field -> original_header (first wins).
    field_to_header: dict[str, str] = {}
    for header, fieldname in mapping.items():
        if fieldname and fieldname not in field_to_header:
            field_to_header[fieldname] = header

    results: list[RowResult] = []
    seen_slugs: dict[str, int] = {}
    seen_skus: dict[str, int] = {}

    for idx, raw_row in enumerate(rows, start=2):  # row 1 = header
        res = RowResult(row_num=idx)
        data: dict[str, Any] = {}
        for fieldname, header in field_to_header.items():
            value, err = _coerce(fieldname, raw_row.get(header, ""))
            if err:
                res.errors.append(f"{fieldname}: {err}")
            elif value is not None:
                data[fieldname] = value

        # Derive slug from name if missing.
        if not data.get("slug") and data.get("name_en"):
            data["slug"] = slugify(data["name_en"])

        # Required fields (for a new product).
        slug = (data.get("slug") or "").strip()
        is_update = slug in existing_slugs
        for req in ("slug", "name_en", "name_bn", "category", "price"):
            if not is_update and (req not in data or data.get(req) in (None, "")):
                res.errors.append(f"{req} is required")
        if data.get("price") is not None and float(data["price"]) < 0:
            res.errors.append("price cannot be negative")

        # Category matching (only if a category cell was given).
        if data.get("category"):
            node = cat_index.resolve(str(data["category"]))
            if node is None:
                res.errors.append(f"category '{data['category']}' not found in the category tree")
            else:
                res.category_id = node.id
                res.category_label = category_path(node, cat_index)
                data["category"] = node.slug  # canonical legacy string

        # In-file duplicate slug.
        if slug:
            if slug in seen_slugs:
                res.errors.append(f"duplicate slug in file (also row {seen_slugs[slug]})")
            else:
                seen_slugs[slug] = idx

        # SKU duplicate detection (warning, never blocks).
        sku = (data.get("sku") or "").strip()
        if sku:
            if sku in seen_skus:
                res.warnings.append(f"SKU '{sku}' repeats in file (row {seen_skus[sku]})")
            else:
                seen_skus[sku] = idx
            owner = existing_sku_to_slug.get(sku)
            if owner and owner != slug:
                res.warnings.append(f"SKU '{sku}' already used by product '{owner}'")

        # Action.
        if res.errors:
            res.action = "skip"
        elif is_update:
            res.action = "update" if on_existing == "update" else "skip"
        else:
            res.action = "create" if on_new == "create" else "skip"

        res.data = data
        results.append(res)

    return results
